"""
BiasharaWatch — NSE Stock Tracker
Scrapes live prices for all NSE-listed stocks from a live NSE data source.
Runs once per trigger (GitHub Actions calls it every 10 min during trading hours).

Sheets:
  - Market      : All NSE stocks, auto-updated each run
  - My Portfolio: Fill in Ticker, Shares Owned, Buy Price — rest is calculated

Weekly email every Friday at 5:00 PM EAT with your portfolio summary.

Requirements:
    pip install requests beautifulsoup4 openpyxl

Environment variables (set as GitHub Secrets):
    GMAIL_ADDRESS   — your Gmail address
    GMAIL_APP_PASS  — Gmail App Password (not your login password)
    NOTIFY_EMAIL    — recipient email for the weekly report
"""

import os
import sys
import json
import argparse
import smtplib
from datetime import datetime, timezone, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Load config ───────────────────────────────────────────────────────────────
CONFIG_FILE = Path(__file__).parent / "config.json"

def load_config() -> dict:
    if CONFIG_FILE.exists():
        with open(CONFIG_FILE) as f:
            return json.load(f)
    # Fallback defaults if config.json is missing
    return {
        "trading_hours": {"start": 9, "end": 17},
        "email":         {"send_on": "Friday", "send_at_hour": 17},
        "output_file":   "stock_prices.xlsx",
        "source_url":    "https://example.com/nse-data",  # set your source URL in config.json
    }

CONFIG      = load_config()
SOURCE_URL  = CONFIG["source_url"]
OUTPUT_FILE = CONFIG["output_file"]
EAT         = timezone(timedelta(hours=3))

# ── Styles ────────────────────────────────────────────────────────────────────
GREEN       = "006633"
WHITE       = "FFFFFF"
LIGHT_GREEN = "E8F5E9"

HEADER_FILL   = PatternFill("solid", start_color=GREEN)
ALT_FILL      = PatternFill("solid", start_color=LIGHT_GREEN)
HEADER_FONT   = Font(name="Arial", bold=True, color=WHITE, size=11)
TICKER_FONT   = Font(name="Arial", bold=True, color=GREEN, size=11)
DATA_FONT     = Font(name="Arial", size=11)
POSITIVE_FONT = Font(name="Arial", size=11, color="1B5E20")
NEGATIVE_FONT = Font(name="Arial", size=11, color="B71C1C")
CENTER        = Alignment(horizontal="center", vertical="center")
LEFT          = Alignment(horizontal="left",   vertical="center")


# ── Scraper ───────────────────────────────────────────────────────────────────

def scrape_prices() -> list:
    """
    Fetch live NSE stock prices from the configured data source.

    The site is a Nuxt.js SPA. All stock data is server-side rendered into a
    <script id="__NUXT_DATA__"> tag as a flat JSON reference array — no
    JavaScript execution required. We parse that directly.

    Data structure:
        nuxt[4]  = {"status": 5, "message": 6, "data": 7, ...}
        nuxt[7]  = [8, 28, 45, ...]           # list of stock object indices
        nuxt[8]  = {"symbol": 10, "company_name": 9, "close": 25, ...}
        nuxt[10] = "ABSA"
        nuxt[25] = 29.5
        nuxt[20] = {"name": 21, ...}           # sector object
        nuxt[21] = "Banking"
    """
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/124.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
    }
    try:
        resp = requests.get(SOURCE_URL, headers=headers, timeout=30)
        resp.raise_for_status()
    except requests.RequestException as e:
        print(f"[ERROR] Could not reach {SOURCE_URL}: {e}", file=sys.stderr)
        return []

    soup = BeautifulSoup(resp.text, "html.parser")
    tag  = soup.find("script", {"id": "__NUXT_DATA__"})
    if not tag:
        print("[ERROR] __NUXT_DATA__ tag not found — site structure may have changed.",
              file=sys.stderr)
        return []

    try:
        nuxt = json.loads(tag.string)
    except Exception as e:
        print(f"[ERROR] Failed to parse __NUXT_DATA__ JSON: {e}", file=sys.stderr)
        return []

    def get(key_str, stock_dict):
        """Dereference a string key from a stock dict via the nuxt array."""
        idx = stock_dict.get(key_str)
        if idx is None:
            return None
        return nuxt[idx] if isinstance(idx, int) and idx < len(nuxt) else None

    def find_stock_list():
        """
        Locate the list of stock-object indices.
        Primary path: nuxt[4]["data"] -> index of the list.
        Fallback: scan for any list of 30+ ints (robust to minor structure shifts).
        """
        entry = nuxt[4] if len(nuxt) > 4 else None
        if isinstance(entry, dict) and "data" in entry:
            candidate_idx = entry["data"]
            if isinstance(candidate_idx, int) and isinstance(nuxt[candidate_idx], list):
                return nuxt[candidate_idx]
        for item in nuxt:
            if (isinstance(item, list) and len(item) >= 30
                    and all(isinstance(x, int) for x in item)):
                return item
        return None

    stock_indices = find_stock_list()
    if not stock_indices:
        print("[ERROR] Could not locate stock list in __NUXT_DATA__.", file=sys.stderr)
        return []

    print(f"  Found {len(stock_indices)} stocks in page data.")
    eat_now = datetime.now(EAT).strftime("%Y-%m-%d %H:%M EAT")
    results = []

    for idx in stock_indices:
        try:
            raw = nuxt[idx]
            if not isinstance(raw, dict):
                continue

            ticker  = (get("symbol",        raw) or "").strip()
            company = (get("company_name",   raw) or "").strip()
            close   =  get("close",          raw)
            prev    =  get("previous_price", raw)
            volume  =  get("volume",         raw)

            # Sector is a nested object: raw["sector"] -> idx -> {"name": idx2, ...}
            sector_name = ""
            sec_idx = raw.get("sector")
            if isinstance(sec_idx, int) and sec_idx < len(nuxt):
                sec_obj = nuxt[sec_idx]
                if isinstance(sec_obj, dict):
                    name_idx = sec_obj.get("name")
                    if isinstance(name_idx, int) and name_idx < len(nuxt):
                        sector_name = str(nuxt[name_idx] or "")

            if not ticker or close is None:
                continue

            try:
                change_pct = round(((float(close) - float(prev)) / float(prev)) * 100, 2) \
                             if prev and float(prev) != 0 else 0.0
            except Exception:
                change_pct = 0.0

            vol = int(volume) if isinstance(volume, (int, float)) else 0

            results.append({
                "ticker":    ticker,
                "company":   company,
                "sector":    sector_name,
                "price":     float(close),
                "change":    change_pct,
                "volume":    vol,
                "timestamp": eat_now,
            })

            sign = "+" if change_pct >= 0 else ""
            print(f"  {ticker:8s}  KES {float(close):>10.2f}  {sign}{change_pct}%")

        except Exception as e:
            print(f"  [WARN] Skipping idx {idx}: {e}", file=sys.stderr)

    return results


# ── Workbook setup ────────────────────────────────────────────────────────────

def setup_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Market"
    _setup_market_headers(ws)
    ps = wb.create_sheet("My Portfolio")
    _setup_portfolio_sheet(ps)
    return wb


def _setup_market_headers(ws):
    headers = ["Ticker", "Company", "Sector", "Price (KES)", "Change (%)", "Volume", "Last Updated"]
    widths  = [10,        28,         22,        14,            12,            14,        22]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def _setup_portfolio_sheet(ps):
    ps.merge_cells("A1:I1")
    title       = ps["A1"]
    title.value = "My NSE Portfolio"
    title.font  = Font(name="Arial", bold=True, size=14, color=GREEN)
    title.alignment = CENTER
    title.fill  = PatternFill("solid", start_color=LIGHT_GREEN)

    ps.merge_cells("A2:I2")
    note       = ps["A2"]
    note.value = "Fill in columns B (Ticker), D (Shares Owned) and E (Buy Price). Columns C and F–I are auto-calculated."
    note.font  = Font(name="Arial", italic=True, size=10, color="555555")
    note.alignment = LEFT

    headers = ["#", "Ticker", "Company", "Shares Owned", "Buy Price (KES)",
               "Current Price (KES)", "Current Value (KES)", "Gain/Loss (KES)", "Gain/Loss (%)"]
    widths  = [5, 10, 28, 14, 15, 18, 18, 18, 14]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ps.cell(row=3, column=col, value=h)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTER
        ps.column_dimensions[get_column_letter(col)].width = w

    for row in range(4, 24):
        ps.cell(row=row, column=1, value=row - 3).font = DATA_FONT
        ps.cell(row=row, column=1).alignment = CENTER
        if row % 2 == 0:
            for col in range(1, 10):
                ps.cell(row=row, column=col).fill = ALT_FILL

    ps.freeze_panes = "A4"
    ps.sheet_view.showGridLines = False


# ── Workbook update ───────────────────────────────────────────────────────────

def update_market_sheet(wb: openpyxl.Workbook, stocks: list):
    ws = wb["Market"]

    # Rewrite headers on every run — corrects any stale column layout from old files
    _setup_market_headers(ws)

    for row in ws.iter_rows(min_row=2, max_row=max(ws.max_row, 2)):
        for cell in row:
            cell.value = None
            cell.font  = DATA_FONT
            cell.fill  = PatternFill()

    for i, d in enumerate(stocks, 1):
        row = i + 1
        cf  = POSITIVE_FONT if d["change"] >= 0 else NEGATIVE_FONT

        ws.cell(row=row, column=1, value=d["ticker"]).font    = TICKER_FONT
        ws.cell(row=row, column=2, value=d["company"]).font   = DATA_FONT
        ws.cell(row=row, column=3, value=d["sector"]).font    = DATA_FONT
        ws.cell(row=row, column=4, value=d["price"]).font     = DATA_FONT
        ws.cell(row=row, column=5, value=d["change"] / 100).font = cf
        ws.cell(row=row, column=6, value=d["volume"]).font    = DATA_FONT
        ws.cell(row=row, column=7, value=d["timestamp"]).font = DATA_FONT

        ws.cell(row=row, column=4).number_format = '#,##0.00'
        ws.cell(row=row, column=5).number_format = '+0.00%;-0.00%'
        ws.cell(row=row, column=6).number_format = '#,##0'

        if row % 2 == 0:
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = ALT_FILL


def update_portfolio_formulas(wb: openpyxl.Workbook):
    ps = wb["My Portfolio"]
    for row in range(4, ps.max_row + 1):
        if not ps.cell(row=row, column=2).value:
            continue
        # Company name — lookup against Market col A (Ticker) : col B (Company)
        ps.cell(row=row, column=3).value = f'=IFERROR(VLOOKUP(B{row},Market!A:B,2,0),"")'
        ps.cell(row=row, column=3).font  = DATA_FONT
        # Current price — lookup against Market col A (Ticker) : col D (Price)
        ps.cell(row=row, column=6).value = f'=IFERROR(VLOOKUP(B{row},Market!A:D,4,0),"")'
        ps.cell(row=row, column=6).number_format = '#,##0.00'
        ps.cell(row=row, column=6).font  = DATA_FONT
        ps.cell(row=row, column=7).value = f'=IFERROR(D{row}*F{row},"")'
        ps.cell(row=row, column=7).number_format = '#,##0.00'
        ps.cell(row=row, column=7).font  = DATA_FONT
        ps.cell(row=row, column=8).value = f'=IFERROR((F{row}-E{row})*D{row},"")'
        ps.cell(row=row, column=8).number_format = '+#,##0.00;-#,##0.00'
        ps.cell(row=row, column=8).font  = DATA_FONT
        ps.cell(row=row, column=9).value = f'=IFERROR((F{row}-E{row})/E{row},"")'
        ps.cell(row=row, column=9).number_format = '+0.00%;-0.00%'
        ps.cell(row=row, column=9).font  = DATA_FONT


# ── Weekly email ──────────────────────────────────────────────────────────────

def send_weekly_email(wb: openpyxl.Workbook, stocks: list):
    gmail_addr  = os.environ.get("GMAIL_ADDRESS")
    gmail_pass  = os.environ.get("GMAIL_APP_PASS")
    notify_addr = os.environ.get("NOTIFY_EMAIL")

    if not all([gmail_addr, gmail_pass, notify_addr]):
        print("  [SKIP] Email secrets not configured — set GMAIL_ADDRESS, GMAIL_APP_PASS, NOTIFY_EMAIL.")
        return

    price_map      = {d["ticker"]: d["price"] for d in stocks}
    ps             = wb["My Portfolio"]
    rows_html      = ""
    total_invested = total_value = 0

    for row in range(4, ps.max_row + 1):
        ticker    = ps.cell(row=row, column=2).value
        shares    = ps.cell(row=row, column=4).value
        buy_price = ps.cell(row=row, column=5).value
        if not ticker or not shares or not buy_price:
            continue
        cur_price = price_map.get(str(ticker).upper(), 0)
        if not cur_price:
            continue
        cur_value       = shares * cur_price
        invested        = shares * buy_price
        gain_kes        = cur_value - invested
        gain_pct        = (gain_kes / invested * 100) if invested else 0
        total_invested += invested
        total_value    += cur_value
        color = "#1B5E20" if gain_kes >= 0 else "#B71C1C"
        sign  = "+" if gain_kes >= 0 else ""
        rows_html += f"""
        <tr style="border-bottom:1px solid #eee">
          <td style="padding:8px">{ticker}</td>
          <td style="padding:8px">{ps.cell(row=row, column=3).value or ticker}</td>
          <td style="padding:8px;text-align:right">{shares:,.0f}</td>
          <td style="padding:8px;text-align:right">KES {buy_price:,.2f}</td>
          <td style="padding:8px;text-align:right">KES {cur_price:,.2f}</td>
          <td style="padding:8px;text-align:right">KES {cur_value:,.2f}</td>
          <td style="padding:8px;text-align:right;color:{color}">{sign}KES {gain_kes:,.2f}</td>
          <td style="padding:8px;text-align:right;color:{color}">{sign}{gain_pct:.2f}%</td>
        </tr>"""

    total_gain     = total_value - total_invested
    total_gain_pct = (total_gain / total_invested * 100) if total_invested else 0
    total_color    = "#1B5E20" if total_gain >= 0 else "#B71C1C"
    total_sign     = "+" if total_gain >= 0 else ""
    today          = datetime.now(EAT).strftime("%A, %d %B %Y")

    html = f"""
    <html><body style="font-family:Arial,sans-serif;color:#222;max-width:820px;margin:auto">
      <div style="background:#006633;color:white;padding:20px;border-radius:8px 8px 0 0">
        <h2 style="margin:0">biasharaWatch — Weekly Portfolio Report</h2>
        <p style="margin:4px 0 0">NSE closing prices · {today}</p>
      </div>
      <div style="background:#f9f9f9;padding:16px;border:1px solid #ddd;margin-bottom:16px">
        <table width="100%" cellspacing="8"><tr>
          <td><strong>Total Invested</strong><br>
              <span style="font-size:18px">KES {total_invested:,.2f}</span></td>
          <td><strong>Current Value</strong><br>
              <span style="font-size:18px">KES {total_value:,.2f}</span></td>
          <td><strong>Total Gain/Loss</strong><br>
              <span style="font-size:18px;color:{total_color}">
                {total_sign}KES {total_gain:,.2f} ({total_sign}{total_gain_pct:.2f}%)
              </span>
          </td>
        </tr></table>
      </div>
      <table width="100%" cellspacing="0" style="border-collapse:collapse;font-size:13px">
        <thead><tr style="background:#006633;color:white">
          <th style="padding:8px;text-align:left">Ticker</th>
          <th style="padding:8px;text-align:left">Company</th>
          <th style="padding:8px;text-align:right">Shares</th>
          <th style="padding:8px;text-align:right">Buy Price</th>
          <th style="padding:8px;text-align:right">Current Price</th>
          <th style="padding:8px;text-align:right">Current Value</th>
          <th style="padding:8px;text-align:right">Gain/Loss</th>
          <th style="padding:8px;text-align:right">Gain/Loss %</th>
        </tr></thead>
        <tbody>{rows_html}</tbody>
      </table>
      <p style="font-size:11px;color:#888;margin-top:16px">
        NSE market data · Generated every Friday at 5:00 PM EAT.
      </p>
    </body></html>"""

    msg            = MIMEMultipart("alternative")
    msg["Subject"] = f"biasharaWatch Weekly Portfolio Report — {datetime.now(EAT).strftime('%d %b %Y')}"
    msg["From"]    = f"biasharaWatch <{gmail_addr}>"
    msg["To"]      = notify_addr
    msg.attach(MIMEText(html, "html"))

    try:
        with smtplib.SMTP("smtp.gmail.com", 587) as server:
            server.starttls()
            server.login(gmail_addr, gmail_pass)
            server.sendmail(gmail_addr, notify_addr, msg.as_string())
        print("  Weekly report sent!")
    except Exception as e:
        print(f"  [ERROR] Email failed: {e}", file=sys.stderr)


# ── Main ──────────────────────────────────────────────────────────────────────

def parse_args():
    parser = argparse.ArgumentParser(description="biasharaWatch — NSE Stock Tracker")
    parser.add_argument(
        "--force", "-f",
        action="store_true",
        help="Run regardless of trading hours (useful for testing)"
    )
    parser.add_argument(
        "--start", type=int, default=None, metavar="HOUR",
        help="Override trading start hour in EAT (e.g. --start 9)"
    )
    parser.add_argument(
        "--end", type=int, default=None, metavar="HOUR",
        help="Override trading end hour in EAT (e.g. --end 17)"
    )
    parser.add_argument(
        "--send-email",
        action="store_true",
        help="Force send the weekly portfolio email right now"
    )
    return parser.parse_args()


def is_trading_hours(now_eat: datetime, start: int, end: int) -> bool:
    day_name     = now_eat.strftime("%A")
    allowed_days = CONFIG["trading_hours"].get(
        "days", ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
    )
    return day_name in allowed_days and start <= now_eat.hour < end


def main():
    args    = parse_args()
    now_eat = datetime.now(EAT)

    start_hour = args.start if args.start is not None else CONFIG["trading_hours"]["start"]
    end_hour   = args.end   if args.end   is not None else CONFIG["trading_hours"]["end"]

    email_cfg   = CONFIG["email"]
    send_day    = email_cfg.get("send_on", "Friday")
    send_hour   = email_cfg.get("send_at_hour", 17)

    in_hours    = is_trading_hours(now_eat, start_hour, end_hour)
    is_send_day = now_eat.strftime("%A") == send_day
    is_closing  = now_eat.hour == send_hour and now_eat.minute < 15

    print("biasharaWatch — NSE Stock Tracker")
    print(f"Time : {now_eat.strftime('%Y-%m-%d %H:%M EAT')} ({now_eat.strftime('%A')})")
    print(f"Hours: {start_hour}:00 – {end_hour}:00 EAT  |  In trading window: {in_hours}")

    if not in_hours and not args.force:
        print("\nOutside trading hours. Use --force to run anyway. Exiting.")
        sys.exit(0)

    if args.force and not in_hours:
        print("--force flag set — running outside trading hours.\n")
    else:
        print()

    out_path = Path(OUTPUT_FILE)
    wb       = openpyxl.load_workbook(out_path) if out_path.exists() else setup_workbook()

    print("Scraping live NSE prices...")
    stocks = scrape_prices()

    if not stocks:
        print("[ERROR] No data retrieved. Exiting without saving.", file=sys.stderr)
        sys.exit(1)

    print(f"\n  Fetched {len(stocks)} stocks.")
    update_market_sheet(wb, stocks)
    update_portfolio_formulas(wb)
    wb.save(out_path)
    print(f"Saved -> {out_path}")

    if args.send_email or (is_send_day and is_closing):
        reason = "--send-email flag" if args.send_email else f"{send_day} {send_hour}:00 EAT"
        print(f"\nSending weekly portfolio report ({reason})...")
        send_weekly_email(wb, stocks)


if __name__ == "__main__":
    main()